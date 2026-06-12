# FILE: scripts/legal_case/_rebuild_xlsx.py
# Purpose: Rebuild a clean, styled delivery log workbook from the original
# Called-by: scripts.legal_case.legal_case_rebuild
# Depends-on: app.styling.themes, app.styling.xlsx_builder
# Last-renovated: 2026-06-11
"""
Rebuild a clean, styled delivery log workbook from the original
spreadsheet data plus the reconciliation mapping.

Sheets:
  - Daily Log  : one row per working day, with an added Evidence Ref column
  - Expenses   : structural copy of the original (empty at present)
  - Dashboard  : KPI summary sheet

The Volume Evidence sheet from the original is dropped (empty template).

Uses ASTRA's own build_xlsx from app.styling.xlsx_builder so the output
picks up the astra_minimal theme automatically via keyword detection.
"""
from __future__ import annotations

import json
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def _format_iso(v: Any) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if v is None:
        return ""
    return str(v)


def _format_time(v: Any) -> str:
    if v is None:
        return ""
    # openpyxl returns datetime.time for time-formatted cells.
    try:
        return v.strftime("%H:%M")
    except Exception:
        return str(v)


def _round_num(v: Any, places: int = 2) -> Any:
    if v is None:
        return None
    try:
        return round(float(v), places)
    except (TypeError, ValueError):
        return v


def rebuild(
    source_xlsx: Path,
    reconciliation_path: Path,
    output_path: Path,
) -> Dict[str, Any]:
    """Read the source workbook + reconciliation mapping; emit a styled
    workbook at output_path. Returns the build_xlsx result dict."""
    import openpyxl

    recon = json.loads(reconciliation_path.read_text(encoding="utf-8"))
    # date -> list of screenshot filenames
    evidence_by_date: Dict[str, List[str]] = {
        m["date"]: m.get("screenshot_filenames") or []
        for m in recon.get("matches", [])
    }

    wb = openpyxl.load_workbook(str(source_xlsx), data_only=True)

    # ---- Daily Log ----
    ws = wb["Daily Log"]
    headers: List[str] = []
    for cell in ws[3]:
        v = cell.value
        headers.append(str(v).replace("\n", " ").strip() if v is not None else "")

    # Drop empty trailing header slots.
    while headers and headers[-1] == "":
        headers.pop()

    daily_rows: List[List[Any]] = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        first = r[0]
        if isinstance(first, str) and first.strip().upper() in {"TOTALS", "AVERAGES", "EFF. HOURLY RATE"}:
            continue
        row = list(r[: len(headers)])
        iso = _format_iso(row[0])
        # Time columns (indices 2..5 in the current schema: Leave/Depot/Last/Home).
        row[0] = iso
        for i in (2, 3, 4, 5):
            if i < len(row):
                row[i] = _format_time(row[i])
        # Round numeric columns to 2 dp for display.
        for i in range(6, len(row)):
            row[i] = _round_num(row[i])

        # Append the Evidence Ref column: comma-separated screenshot filenames.
        refs = evidence_by_date.get(iso, [])
        row.append(", ".join(refs) if refs else "no finished-day screenshot")
        daily_rows.append(row)

    daily_headers = headers + ["Evidence Ref"]

    # ---- Expenses (structural only; sheet is empty in the source) ----
    exp_ws = wb["Expenses"]
    exp_headers: List[str] = []
    for cell in exp_ws[2]:
        v = cell.value
        exp_headers.append(str(v).replace("\n", " ").strip() if v is not None else "")
    while exp_headers and exp_headers[-1] == "":
        exp_headers.pop()

    exp_rows: List[List[Any]] = []
    for r in exp_ws.iter_rows(min_row=3, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        first = r[0]
        # Skip the categories-legend line and the TOTAL marker line.
        if isinstance(first, str) and (first.strip().upper().startswith("CATEGOR") or first.strip().upper() == "TOTAL"):
            continue
        exp_rows.append([_format_iso(r[0])] + list(r[1:len(exp_headers)]))

    # ---- Dashboard (rebuild as a flat KPI table) ----
    dash_ws = wb["Dashboard"]
    kpi_rows: List[List[Any]] = []
    for r in dash_ws.iter_rows(min_row=3, values_only=True):
        # Dashboard layout is columns B, C, and E, F with key/value pairs.
        # Flatten into a simple two-column KPI list.
        if len(r) >= 3 and r[1] and r[2] is not None:
            kpi_rows.append([str(r[1]).strip(), _round_num(r[2])])
        if len(r) >= 6 and r[4] and r[5] is not None:
            kpi_rows.append([str(r[4]).strip(), _round_num(r[5])])

    sheets = [
        {
            "name": "Daily Log",
            "headers": daily_headers,
            "rows": daily_rows,
            "freeze_header": True,
            "auto_filter": True,
        },
        {
            "name": "Expenses",
            "headers": exp_headers,
            "rows": exp_rows,
            "freeze_header": True,
            "auto_filter": True,
        },
        {
            "name": "Dashboard",
            "headers": ["KPI", "Value"],
            "rows": kpi_rows,
            "freeze_header": True,
            "auto_filter": False,
            "column_widths": [45, 20],
        },
    ]

    from app.styling.themes import pick_theme
    from app.styling.xlsx_builder import build_xlsx

    theme = pick_theme(f"delivery_log legal evidence {output_path.name}", "auto")
    result = build_xlsx(str(output_path), sheets, theme, title="Delivery Log — Legal Evidence")
    logger.info(
        "[rebuild_xlsx] Wrote %s (%d rows daily, %d rows expenses, %d KPIs, theme=%s)",
        output_path, len(daily_rows), len(exp_rows), len(kpi_rows), result.get("theme"),
    )
    return result
