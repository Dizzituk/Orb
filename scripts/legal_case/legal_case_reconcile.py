# FILE: scripts/legal_case/legal_case_reconcile.py
# Purpose: Legal case reconciliation.
# Called-by: no static importers found (dynamic/registry use possible)
# Depends-on: scripts.legal_case._bootstrap
# Last-renovated: 2026-06-11
"""
Legal case reconciliation.

Cross-checks the Gemini vision extraction results
(screenshots_ocr.json) against the Daily Log sheet of
delivery_log.xlsx, and produces a reconciliation report that:

- Maps each finished-day screenshot to its Daily Log row by date
- Flags any numeric mismatch between what the screenshot shows
  and what the spreadsheet records
- Lists dates that have spreadsheet entries but no screenshot evidence
- Lists screenshots that couldn't be matched to any date

Output: C:\\Users\\dizzi\\OneDrive\\Documents\\Work Legal\\screenshots_reconciliation.json

Nothing in the original spreadsheet is modified here. This is a pure
cross-check; the rebuild step uses this report as input.
"""
from __future__ import annotations

import json
import logging
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from scripts.legal_case._bootstrap import bootstrap_astra

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

LEGAL_ROOT = Path(r"C:\Users\dizzi\OneDrive\Documents\Work Legal")
OCR_FILE = LEGAL_ROOT / "screenshots_ocr.json"
XLSX_FILE = LEGAL_ROOT / "delivery_log.xlsx"
OUTPUT_FILE = LEGAL_ROOT / "screenshots_reconciliation.json"

# Tolerances. Stop/parcel counts should match exactly; hours within ~15 min.
STOP_TOLERANCE = 0
PARCEL_TOLERANCE = 0
HOURS_TOLERANCE = 0.25


# =============================================================================
# DATE NORMALISATION
# =============================================================================

_MONTHS = {
    "january": 1, "jan": 1, "february": 2, "feb": 2, "march": 3, "mar": 3,
    "april": 4, "apr": 4, "may": 5, "june": 6, "jun": 6, "july": 7, "jul": 7,
    "august": 8, "aug": 8, "september": 9, "sep": 9, "sept": 9,
    "october": 10, "oct": 10, "november": 11, "nov": 11, "december": 12, "dec": 12,
}

_VISIBLE_RE = re.compile(
    r"(?:(?P<d1>\d{1,2})(?:st|nd|rd|th)?\s+(?P<mon1>[A-Za-z]+))"  # "18th November"
    r"|(?:(?P<mon2>[A-Za-z]+)\s+(?P<d2>\d{1,2})(?:st|nd|rd|th)?)"  # "November 18"
    r"|(?:(?P<d3>\d{1,2})[\-/](?P<m3>\d{1,2})[\-/](?P<y3>\d{2,4}))"  # "18/11/2025"
    r"|(?:(?P<y4>\d{4})[\-/](?P<m4>\d{1,2})[\-/](?P<d4>\d{1,2}))"   # "2025-11-18"
)


def _normalise_visible_date(visible: Optional[str], filename_date: Optional[str]) -> Optional[str]:
    """Return a canonical YYYY-MM-DD string, preferring the filename date
    (which includes the year) as the source of truth when a visible date
    is ambiguous (year-less)."""
    # Filename date always wins when we have one, because screenshot text
    # is usually year-less ("18th November").
    if filename_date:
        return filename_date

    if not visible:
        return None

    m = _VISIBLE_RE.search(visible.strip())
    if not m:
        return None

    try:
        if m.group("d1") and m.group("mon1"):
            mon = _MONTHS.get(m.group("mon1").lower())
            if not mon:
                return None
            return f"????-{mon:02d}-{int(m.group('d1')):02d}"
        if m.group("mon2") and m.group("d2"):
            mon = _MONTHS.get(m.group("mon2").lower())
            if not mon:
                return None
            return f"????-{mon:02d}-{int(m.group('d2')):02d}"
        if m.group("d3") and m.group("m3") and m.group("y3"):
            y = int(m.group("y3"))
            if y < 100:
                y += 2000
            return f"{y:04d}-{int(m.group('m3')):02d}-{int(m.group('d3')):02d}"
        if m.group("y4"):
            return f"{int(m.group('y4')):04d}-{int(m.group('m4')):02d}-{int(m.group('d4')):02d}"
    except Exception:
        return None
    return None


# =============================================================================
# SPREADSHEET LOAD
# =============================================================================

def _load_daily_log(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Read the Daily Log sheet into a list of row dicts keyed by header."""
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    if "Daily Log" not in wb.sheetnames:
        raise ValueError("Daily Log sheet not found in workbook")
    ws = wb["Daily Log"]

    # Header row is row 3 (row 1 = title, row 2 = instruction note).
    headers: List[str] = []
    for cell in ws[3]:
        v = cell.value
        headers.append(str(v).replace("\n", " ").strip() if v is not None else "")

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        # Skip completely blank rows.
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        first = r[0]
        # Bottom of sheet has TOTALS / AVERAGES / EFF. HOURLY RATE marker rows.
        if isinstance(first, str) and first.strip().upper() in {"TOTALS", "AVERAGES", "EFF. HOURLY RATE"}:
            continue
        rec: Dict[str, Any] = {}
        for h, v in zip(headers, r):
            if h:
                rec[h] = v
        rows.append(rec)
    return rows


def _row_iso_date(row: Dict[str, Any]) -> Optional[str]:
    """Pull the row's Date field and format as YYYY-MM-DD."""
    v = row.get("Date")
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str) and v.strip():
        try:
            return datetime.fromisoformat(v.strip()).date().isoformat()
        except Exception:
            return None
    return None


# =============================================================================
# COMPARISON
# =============================================================================

def _compare(row: Dict[str, Any], shot: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Return a list of mismatches between spreadsheet row and screenshot."""
    mismatches: List[Dict[str, Any]] = []
    fields = (shot.get("extraction") or {}).get("fields") or {}

    def _check(label: str, sheet_key: str, shot_key: str, tol: float) -> None:
        sv = row.get(sheet_key)
        xv = fields.get(shot_key)
        if sv is None or xv is None:
            return
        try:
            if abs(float(sv) - float(xv)) > tol:
                mismatches.append({
                    "field": label,
                    "spreadsheet": sv,
                    "screenshot": xv,
                    "delta": float(sv) - float(xv),
                })
        except (TypeError, ValueError):
            return

    # Parcels and failed counts SHOULD match exactly - both measure the
    # same concept (parcels delivered, parcels failed) and any delta here
    # is a real discrepancy worth flagging.
    _check("parcels", "Parcels Delivered", "parcels", PARCEL_TOLERANCE)
    _check("failed", "Failed Deliveries", "failed", 0)

    # Hours within ~15 min tolerance (the screenshot shows a rounded
    # "Duration" that the spreadsheet computes from start/end timestamps).
    _check("hours", "Delivery Time (hrs)", "hours_worked", HOURS_TOLERANCE)

    # NOTE: stops are intentionally NOT compared. "Stops Assigned" on the
    # spreadsheet is the manifest count (everything on the route). The
    # screenshot's "stops_delivered" is the count of unique stops
    # physically reached. These legitimately differ when some stops are
    # carried over, removed by management, or combined at a single
    # address. Both numbers are kept separately in the output row.
    return mismatches


# =============================================================================
# MAIN
# =============================================================================

def run() -> Dict[str, Any]:
    bootstrap_astra()

    if not OCR_FILE.exists():
        raise FileNotFoundError(f"Missing {OCR_FILE} — run legal_case_extractor first.")
    ocr = json.loads(OCR_FILE.read_text(encoding="utf-8"))

    # Build screenshot index: iso_date -> list of finished-day screenshots.
    shots_by_date: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    unmatched_shots: List[Dict[str, Any]] = []

    for item in ocr.get("items", []):
        ext = item.get("extraction") or {}
        if ext.get("type") != "route_summary_finished":
            continue
        iso = _normalise_visible_date(ext.get("date_visible"), item.get("filename_date"))
        if iso and not iso.startswith("????"):
            shots_by_date[iso].append(item)
        else:
            unmatched_shots.append({
                "filename": item.get("filename"),
                "reason": "could not normalise date",
                "visible": ext.get("date_visible"),
            })

    # Load spreadsheet rows.
    rows = _load_daily_log(XLSX_FILE)
    logger.info("[reconcile] Loaded %d Daily Log rows", len(rows))
    logger.info("[reconcile] %d dates have screenshot evidence", len(shots_by_date))

    matches: List[Dict[str, Any]] = []
    dates_missing_evidence: List[str] = []

    for row in rows:
        iso = _row_iso_date(row)
        if not iso:
            continue
        shots = shots_by_date.get(iso, [])
        if not shots:
            # Only flag as missing if the row actually has work logged.
            if row.get("Stops Assigned") or row.get("Parcels Delivered"):
                dates_missing_evidence.append(iso)
            continue

        # Use the first shot as primary; list all filenames in evidence.
        primary = shots[0]
        mismatches = _compare(row, primary)
        matches.append({
            "date": iso,
            "day": row.get("Day"),
            "screenshot_filenames": [s["filename"] for s in shots],
            "primary_screenshot": primary["filename"],
            "sheet_stops": row.get("Stops Assigned"),
            "sheet_parcels": row.get("Parcels Delivered"),
            "sheet_failed": row.get("Failed Deliveries"),
            "sheet_delivery_hours": row.get("Delivery Time (hrs)"),
            "sheet_total_hours": row.get("Total Hours Worked"),
            "shot_stops": (primary.get("extraction") or {}).get("fields", {}).get("stops_delivered"),
            "shot_parcels": (primary.get("extraction") or {}).get("fields", {}).get("parcels"),
            "shot_failed": (primary.get("extraction") or {}).get("fields", {}).get("failed"),
            "shot_hours": (primary.get("extraction") or {}).get("fields", {}).get("hours_worked"),
            "route_name": (primary.get("extraction") or {}).get("fields", {}).get("route_name"),
            "mismatches": mismatches,
            "status": "mismatch" if mismatches else "match",
        })

    matches.sort(key=lambda x: x["date"])
    dates_missing_evidence.sort()

    n_match = sum(1 for m in matches if m["status"] == "match")
    n_mismatch = sum(1 for m in matches if m["status"] == "mismatch")

    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": {
            "spreadsheet_rows": len(rows),
            "screenshot_dates_matched": len(matches),
            "exact_matches": n_match,
            "mismatches": n_mismatch,
            "dates_without_evidence": len(dates_missing_evidence),
            "screenshots_without_date": len(unmatched_shots),
        },
        "matches": matches,
        "dates_without_evidence": dates_missing_evidence,
        "screenshots_without_date": unmatched_shots,
    }
    OUTPUT_FILE.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")

    logger.info("[reconcile] Summary: %s", payload["summary"])
    logger.info("[reconcile] Wrote %s", OUTPUT_FILE)
    return payload


if __name__ == "__main__":
    run()
