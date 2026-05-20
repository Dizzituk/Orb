# FILE: app/styling/xlsx_builder.py
"""
Build styled Excel workbooks from a structured sheet list.

Sheets is a list of:
  {
    "name": "Sheet1",
    "headers": ["Col A", "Col B", "Col C"],
    "rows": [["a1", "b1", "c1"], ["a2", "b2", "c2"]],
    "freeze_header": True,        # default True
    "auto_filter": True,          # default True
    "column_widths": [12, 30, 18] # optional, auto-fits if omitted
  }

Header row gets theme fill colour, white bold text. Body rows get optional
zebra striping. Column widths auto-fit to content (capped at 60 chars) if
not specified.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List


def build_xlsx(
    output_path: str,
    sheets: List[Dict[str, Any]],
    theme: Dict[str, Any],
    title: str = "",
) -> Dict[str, Any]:
    """Render a styled .xlsx workbook. Returns {path, size_bytes, sheets_rendered}."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Remove the default empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    header_fill_hex = theme.get("table_header_fill", "0B5FFF").lstrip("#")
    header_text_hex = theme.get("table_header_text", "FFFFFF").lstrip("#")
    zebra_hex = theme.get("colour_zebra", "F8FAFC").lstrip("#")
    rule_hex = theme.get("colour_rule", "E5E7EB").lstrip("#")
    body_font_name = theme.get("font_body_single", "Calibri")
    body_size = theme.get("body_size_pt", 11)
    use_zebra = theme.get("table_zebra", True)

    header_fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
    header_font = Font(name=body_font_name, size=body_size, bold=True, color=header_text_hex)
    body_font = Font(name=body_font_name, size=body_size, color=theme.get("colour_text", "1A1F2E").lstrip("#"))
    zebra_fill = PatternFill(start_color=zebra_hex, end_color=zebra_hex, fill_type="solid")
    thin = Side(border_style="thin", color=rule_hex)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheets_rendered = 0

    for sheet_def in sheets:
        try:
            name = (sheet_def.get("name") or f"Sheet{sheets_rendered + 1}")[:31]  # Excel limit
            ws = wb.create_sheet(title=name)

            headers = sheet_def.get("headers", [])
            rows = sheet_def.get("rows", [])
            freeze = sheet_def.get("freeze_header", True)
            do_filter = sheet_def.get("auto_filter", True)
            widths = sheet_def.get("column_widths")

            # Header row
            if headers:
                for ci, h in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=ci, value=str(h))
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = cell_border

            # Body rows
            row_offset = 2 if headers else 1
            for ri, row in enumerate(rows):
                excel_row = ri + row_offset
                for ci, val in enumerate(row, start=1):
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.font = body_font
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = cell_border
                    if use_zebra and ri % 2 == 1:
                        cell.fill = zebra_fill

            # Column widths
            n_cols = max(len(headers), max((len(r) for r in rows), default=0))
            if widths:
                for i, w in enumerate(widths[:n_cols], start=1):
                    ws.column_dimensions[get_column_letter(i)].width = float(w)
            else:
                for i in range(1, n_cols + 1):
                    max_len = 10  # minimum
                    if headers and i <= len(headers):
                        max_len = max(max_len, len(str(headers[i - 1])))
                    for row in rows:
                        if i <= len(row):
                            cell_str = str(row[i - 1] if row[i - 1] is not None else "")
                            max_len = max(max_len, min(60, len(cell_str)))
                    ws.column_dimensions[get_column_letter(i)].width = max_len + 2

            # Freeze + filter
            if freeze and headers:
                ws.freeze_panes = "A2"
            if do_filter and headers and rows:
                last_col = get_column_letter(len(headers))
                ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

            sheets_rendered += 1
        except Exception:
            continue

    # Make sure we have at least one sheet (Excel requires it)
    if sheets_rendered == 0:
        ws = wb.create_sheet(title="Sheet1")
        ws["A1"] = "(empty workbook)"

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    size = out.stat().st_size

    return {
        "path": str(out),
        "size_bytes": size,
        "sheets_rendered": sheets_rendered,
        "theme": theme.get("name", "unknown"),
    }
