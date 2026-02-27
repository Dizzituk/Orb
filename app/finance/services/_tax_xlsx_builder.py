# FILE: app/finance/services/_tax_xlsx_builder.py
"""
Builds the multi-tab Excel workbook for the tax pack.

Called by tax_export_service.generate_tax_xlsx().
Writes to a BytesIO buffer — no disk I/O.
"""
from __future__ import annotations

import io
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──
NAVY_FILL = PatternFill('solid', fgColor='1B2A4A')
LIGHT_FILL = PatternFill('solid', fgColor='F1F5F9')
ALT_FILL = PatternFill('solid', fgColor='F8FAFC')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')

HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='Arial', bold=True, color='1B2A4A', size=14)
H2_FONT = Font(name='Arial', bold=True, color='1B2A4A', size=11)
BODY_FONT = Font(name='Arial', size=9.5)
BOLD_FONT = Font(name='Arial', bold=True, size=9.5)
MONEY_FONT = Font(name='Arial', size=9.5)
BLUE_INPUT = Font(name='Arial', size=9.5, color='0000FF')
GREY_FONT = Font(name='Arial', size=8, color='64748B')

GBP_FMT = '£#,##0.00'
INT_FMT = '#,##0'


def _header_row(ws, row: int, headers: list[str]):
    """Apply navy header styling to a row."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _alt_rows(ws, start: int, end: int, max_col: int):
    """Alternate row shading."""
    for r in range(start, end + 1):
        fill = ALT_FILL if (r - start) % 2 == 1 else WHITE_FILL
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = fill


def _set_widths(ws, widths: list[int]):
    """Set column widths."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_tax_xlsx(
    buf: io.BytesIO,
    tax: dict,
    work_logs: list[dict],
    transactions: list[dict],
    cc_transactions: list[dict],
    van: Optional[dict],
    tax_year: str,
) -> None:
    """Build and write the XLSX workbook into buf."""
    wb = Workbook()
    is_actual = tax.get("cost_method") == "actual_costs"
    vc = tax.get("vehicle_costs") or {}

    # ─── Tab 1: Tax Summary ───
    ws = wb.active
    ws.title = "Tax Summary"
    ws.sheet_properties.tabColor = "1B2A4A"

    ws.merge_cells('A1:D1')
    ws['A1'] = f"Self-Employment Tax Summary — {tax_year}"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Generated: {date.today().strftime('%d %B %Y')} | Method: {'Actual Costs' if is_actual else 'Mileage'}"
    ws['A2'].font = GREY_FONT

    _header_row(ws, 4, ["Description", "Amount", "SA103S Box", "Notes"])
    r = 5

    summary_lines = [
        ("Gross Income (Turnover)", tax.get("gross_income", 0), "Box 9/15", "Yodel delivery income"),
        ("Other Business Expenses", tax.get("recorded_expenses", 0), "", "Non-vehicle deductible"),
    ]
    if is_actual and vc:
        summary_lines.append(("Vehicle Running Costs", vc.get("total_running_costs", 0), "Box 20", ""))
        if vc.get("aia", 0) > 0:
            summary_lines.append(("Annual Investment Allowance", vc.get("aia", 0), "Box 49", "Van purchase"))
    if tax.get("home_office_total", 0) > 0:
        summary_lines.append(("Use of Home as Office", tax.get("home_office_total", 0), "Box 25", ""))
    summary_lines.append(("Total Allowable Expenses", tax.get("total_allowable_expenses", 0), "Box 10/20", ""))
    summary_lines.append(("", None, "", ""))
    summary_lines.append(("Taxable Profit", tax.get("taxable_profit", 0), "Box 31", ""))
    summary_lines.append(("Personal Allowance", tax.get("personal_allowance_used", 12570), "", ""))
    summary_lines.append(("", None, "", ""))
    summary_lines.append(("Income Tax", tax.get("total_income_tax", 0), "", ""))
    summary_lines.append(("NI Class 2", tax.get("ni_class2", 0), "", ""))
    summary_lines.append(("NI Class 4", tax.get("ni_class4_main", 0) + tax.get("ni_class4_additional", 0), "", ""))
    summary_lines.append(("TOTAL TAX LIABILITY", tax.get("total_tax_liability", 0), "",
                           f"Effective: {tax.get('effective_tax_rate', 0)}%"))

    for desc, amt, box, note in summary_lines:
        ws.cell(row=r, column=1, value=desc).font = BOLD_FONT if "TOTAL" in desc or desc == "Taxable Profit" else BODY_FONT
        if amt is not None:
            cell = ws.cell(row=r, column=2, value=amt)
            cell.number_format = GBP_FMT
            cell.font = BOLD_FONT if "TOTAL" in desc else MONEY_FONT
        ws.cell(row=r, column=3, value=box).font = BODY_FONT
        ws.cell(row=r, column=4, value=note).font = GREY_FONT
        r += 1

    _set_widths(ws, [35, 18, 14, 32])

    # ─── Tab 2: Work Logs ───
    ws2 = wb.create_sheet("Work Logs")
    ws2.sheet_properties.tabColor = "3B82F6"

    ws2['A1'] = "Daily Work Logs"
    ws2['A1'].font = TITLE_FONT
    ws2['A2'] = f"{len(work_logs)} days recorded"
    ws2['A2'].font = GREY_FONT

    _header_row(ws2, 4, ["Date", "Hours", "Deliveries", "Gross (GBP)", "Per Hour", "Route"])
    for i, log in enumerate(work_logs):
        r = 5 + i
        ws2.cell(row=r, column=1, value=log["date"]).font = BODY_FONT
        ws2.cell(row=r, column=2, value=log["hours"]).font = BODY_FONT
        ws2.cell(row=r, column=2).number_format = '0.0'
        ws2.cell(row=r, column=3, value=log["deliveries"]).font = BODY_FONT
        ws2.cell(row=r, column=3).number_format = INT_FMT
        ws2.cell(row=r, column=4, value=log["gross"]).font = MONEY_FONT
        ws2.cell(row=r, column=4).number_format = GBP_FMT
        ws2.cell(row=r, column=5, value=log["per_hour"]).font = MONEY_FONT
        ws2.cell(row=r, column=5).number_format = GBP_FMT
        ws2.cell(row=r, column=6, value=log["route"]).font = BODY_FONT

    last = 4 + len(work_logs)
    if work_logs:
        _alt_rows(ws2, 5, last, 6)
        tr = last + 1
        ws2.cell(row=tr, column=1, value="TOTALS").font = BOLD_FONT
        ws2.cell(row=tr, column=2, value=f"=SUM(B5:B{last})").font = BOLD_FONT
        ws2.cell(row=tr, column=2).number_format = '0.0'
        ws2.cell(row=tr, column=3, value=f"=SUM(C5:C{last})").font = BOLD_FONT
        ws2.cell(row=tr, column=3).number_format = INT_FMT
        ws2.cell(row=tr, column=4, value=f"=SUM(D5:D{last})").font = BOLD_FONT
        ws2.cell(row=tr, column=4).number_format = GBP_FMT
        ws2.cell(row=tr, column=5, value=f"=D{tr}/B{tr}").font = BOLD_FONT
        ws2.cell(row=tr, column=5).number_format = GBP_FMT
        for c in range(1, 7):
            ws2.cell(row=tr, column=c).fill = LIGHT_FILL

    _set_widths(ws2, [14, 10, 12, 14, 12, 20])

    # ─── Tab 3: Expenses ───
    ws3 = wb.create_sheet("Expenses")
    ws3.sheet_properties.tabColor = "DC2626"

    ws3['A1'] = "All Expense Transactions"
    ws3['A1'].font = TITLE_FONT

    _header_row(ws3, 3, ["Date", "Amount", "Description", "Category", "HMRC Class", "Scope", "Deductible", "Source"])
    for i, tx in enumerate(transactions):
        r = 4 + i
        ws3.cell(row=r, column=1, value=tx["date"]).font = BODY_FONT
        ws3.cell(row=r, column=2, value=tx["amount"]).font = MONEY_FONT
        ws3.cell(row=r, column=2).number_format = GBP_FMT
        ws3.cell(row=r, column=3, value=tx["description"][:60]).font = BODY_FONT
        ws3.cell(row=r, column=4, value=tx["category"]).font = BODY_FONT
        ws3.cell(row=r, column=5, value=tx["hmrc"]).font = BODY_FONT
        ws3.cell(row=r, column=6, value=tx["scope"]).font = BODY_FONT
        ws3.cell(row=r, column=7, value=tx["deductible"]).font = MONEY_FONT
        ws3.cell(row=r, column=7).number_format = GBP_FMT
        ws3.cell(row=r, column=8, value=tx["source"]).font = GREY_FONT

    if transactions:
        _alt_rows(ws3, 4, 3 + len(transactions), 8)

    _set_widths(ws3, [12, 12, 40, 22, 28, 12, 12, 12])

    # ─── Tab 4: Credit Cards ───
    ws4 = wb.create_sheet("Credit Cards")
    ws4.sheet_properties.tabColor = "D97706"

    ws4['A1'] = "Credit Card Transactions"
    ws4['A1'].font = TITLE_FONT

    _header_row(ws4, 3, ["Date", "Amount", "Description", "Card", "Category", "Scope", "Merchant"])
    for i, tx in enumerate(cc_transactions):
        r = 4 + i
        ws4.cell(row=r, column=1, value=tx["date"]).font = BODY_FONT
        ws4.cell(row=r, column=2, value=tx["amount"]).font = MONEY_FONT
        ws4.cell(row=r, column=2).number_format = GBP_FMT
        ws4.cell(row=r, column=3, value=tx["description"][:50]).font = BODY_FONT
        ws4.cell(row=r, column=4, value=tx["card"]).font = BODY_FONT
        ws4.cell(row=r, column=5, value=tx["category"]).font = BODY_FONT
        ws4.cell(row=r, column=6, value=tx["scope"]).font = BODY_FONT
        ws4.cell(row=r, column=7, value=tx["merchant"]).font = BODY_FONT

    if cc_transactions:
        _alt_rows(ws4, 4, 3 + len(cc_transactions), 7)

    _set_widths(ws4, [12, 12, 35, 16, 22, 12, 20])

    # ─── Tab 5: Vehicle (if actual costs) ───
    if is_actual and vc:
        ws5 = wb.create_sheet("Vehicle Costs")
        ws5.sheet_properties.tabColor = "7C3AED"

        ws5['A1'] = "Vehicle Running Costs — Actual Method"
        ws5['A1'].font = TITLE_FONT
        if van:
            ws5['A2'] = van.get("description", "")
            ws5['A2'].font = GREY_FONT

        _header_row(ws5, 4, ["Expense Category", "Amount"])
        vc_lines = [
            ("Fuel", vc.get("fuel", 0)),
            ("HP Interest (deductible)", vc.get("hp_interest", 0)),
            ("HP Capital (NOT deductible)", vc.get("hp_capital", 0)),
            ("Insurance", vc.get("insurance", 0)),
            ("Road Tax (DVLA)", vc.get("road_tax", 0)),
            ("MOT", vc.get("mot", 0)),
            ("Repairs & Tyres", vc.get("repairs", 0)),
            ("Servicing", vc.get("servicing", 0)),
            ("Other Vehicle", vc.get("other_vehicle", 0)),
        ]
        r = 5
        for label, val in vc_lines:
            ws5.cell(row=r, column=1, value=label).font = BODY_FONT
            ws5.cell(row=r, column=2, value=val).font = MONEY_FONT
            ws5.cell(row=r, column=2).number_format = GBP_FMT
            r += 1

        ws5.cell(row=r, column=1, value="Running Total").font = BOLD_FONT
        ws5.cell(row=r, column=2, value=vc.get("total_running_costs", 0)).font = BOLD_FONT
        ws5.cell(row=r, column=2).number_format = GBP_FMT
        ws5.cell(row=r, column=1).fill = LIGHT_FILL
        ws5.cell(row=r, column=2).fill = LIGHT_FILL
        r += 1
        if vc.get("aia", 0) > 0:
            ws5.cell(row=r, column=1, value="Annual Investment Allowance (Van)").font = BOLD_FONT
            ws5.cell(row=r, column=2, value=vc.get("aia", 0)).font = BOLD_FONT
            ws5.cell(row=r, column=2).number_format = GBP_FMT
            r += 1
        ws5.cell(row=r, column=1, value="TOTAL VEHICLE DEDUCTION").font = BOLD_FONT
        ws5.cell(row=r, column=2, value=vc.get("total_deductible", 0)).font = BOLD_FONT
        ws5.cell(row=r, column=2).number_format = GBP_FMT
        ws5.cell(row=r, column=1).fill = LIGHT_FILL
        ws5.cell(row=r, column=2).fill = LIGHT_FILL

        _set_widths(ws5, [35, 18])

    wb.save(buf)
