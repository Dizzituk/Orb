# FILE: app/documents/xlsx_convert.py
# Purpose: xlsx ⇄ Univer workbook snapshot (openpyxl): values, formulas as
#          strings, basic number formats, merged cells. NOT styles-perfect.
# Called-by: app.documents.router
# Depends-on: openpyxl
# Last-renovated: 2026-06-12
"""
xlsx conversion.

WHAT SURVIVES the round trip (v1, by design):
  - cell values (numbers, strings, booleans) and formulas (as strings —
    Univer's formula engine recalculates live)
  - merged cells
  - number-format PATTERNS (e.g. "0.00%", date formats) via Univer styles
WHAT IS DROPPED (documented loss):
  - fonts, colours, borders, column widths/row heights, charts, images,
    pivot tables, data validation, conditional formatting, defined names.
Reads load the workbook twice (formulas + cached values) so formula cells
keep a display value until Univer recalculates.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# Sanity guards — log + truncate instead of shipping megasnapshots to the UI.
MAX_ROWS = 5000
MAX_COLS = 200


def _cell_type(value) -> Optional[int]:
    """Univer CellValueType: 1 string, 2 number, 3 boolean."""
    if isinstance(value, bool):
        return 3
    if isinstance(value, (int, float)):
        return 2
    if isinstance(value, str):
        return 1
    return None


def xlsx_to_snapshot(path: str) -> dict:
    """openpyxl read -> Univer IWorkbookData-shaped dict."""
    from openpyxl import load_workbook

    wb_formulas = load_workbook(path, data_only=False, read_only=True)
    wb_values = load_workbook(path, data_only=True, read_only=True)

    sheets: dict = {}
    sheet_order: list = []
    styles: dict = {}
    style_ids: dict = {}  # pattern -> style id

    for index, ws_f in enumerate(wb_formulas.worksheets):
        ws_v = wb_values.worksheets[index]
        sheet_id = f"sheet-{index + 1:02d}"
        sheet_order.append(sheet_id)

        cell_data: dict = {}
        max_row = 0
        max_col = 0
        truncated = False
        for row in ws_f.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                r, c = cell.row - 1, cell.column - 1
                if r >= MAX_ROWS or c >= MAX_COLS:
                    truncated = True
                    continue
                entry: dict = {}
                raw = cell.value
                if isinstance(raw, str) and raw.startswith("="):
                    entry["f"] = raw
                    cached = ws_v.cell(row=cell.row, column=cell.column).value
                    if cached is not None:
                        entry["v"] = cached if not hasattr(cached, "isoformat") else str(cached)
                        t = _cell_type(entry.get("v"))
                        if t:
                            entry["t"] = t
                else:
                    value = raw if not hasattr(raw, "isoformat") else str(raw)
                    entry["v"] = value
                    t = _cell_type(value)
                    if t:
                        entry["t"] = t
                fmt = getattr(cell, "number_format", None)
                if fmt and fmt != "General":
                    sid = style_ids.get(fmt)
                    if sid is None:
                        sid = f"s{len(style_ids) + 1}"
                        style_ids[fmt] = sid
                        styles[sid] = {"n": {"pattern": fmt}}
                    entry["s"] = sid
                cell_data.setdefault(str(r), {})[str(c)] = entry
                max_row = max(max_row, r)
                max_col = max(max_col, c)
        if truncated:
            logger.warning("[documents] %s sheet %s truncated at %dx%d",
                           path, ws_f.title, MAX_ROWS, MAX_COLS)

        sheets[sheet_id] = {
            "id": sheet_id,
            "name": ws_f.title,
            "rowCount": max(max_row + 50, 100),
            "columnCount": max(max_col + 5, 26),
            "cellData": cell_data,
            # filled by the non-read-only merge pass below (read_only mode
            # doesn't expose merged_cells reliably)
            "mergeData": [],
        }

    # merges need a non-read-only pass (read_only mode hides merged_cells)
    try:
        wb_merge = load_workbook(path, data_only=True, read_only=False)
        for index, ws in enumerate(wb_merge.worksheets):
            if index >= len(sheet_order):
                break
            merge_data = [
                {
                    "startRow": rng.min_row - 1, "endRow": rng.max_row - 1,
                    "startColumn": rng.min_col - 1, "endColumn": rng.max_col - 1,
                }
                for rng in ws.merged_cells.ranges
            ]
            sheets[sheet_order[index]]["mergeData"] = merge_data
        wb_merge.close()
    except Exception as exc:
        logger.warning("[documents] merge extraction failed for %s: %s", path, exc)

    wb_formulas.close()
    wb_values.close()

    return {
        "id": f"wb-{Path(path).stem[:40]}",
        "name": Path(path).name,
        "appVersion": "0.25.0",
        "locale": "enUS",
        "sheetOrder": sheet_order,
        "sheets": sheets,
        "styles": styles,
    }


def snapshot_to_xlsx(snapshot: dict, path: str) -> None:
    """Univer workbook snapshot -> .xlsx via openpyxl (same loss policy)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    styles = snapshot.get("styles") or {}
    order = snapshot.get("sheetOrder") or list((snapshot.get("sheets") or {}).keys())
    for sheet_id in order:
        sheet = (snapshot.get("sheets") or {}).get(sheet_id)
        if not sheet:
            continue
        ws = wb.create_sheet(title=str(sheet.get("name") or sheet_id)[:31])
        for row_key, columns in (sheet.get("cellData") or {}).items():
            for col_key, entry in (columns or {}).items():
                if entry is None:
                    continue
                r, c = int(row_key) + 1, int(col_key) + 1
                cell = ws.cell(row=r, column=c)
                formula = entry.get("f")
                if formula:
                    cell.value = formula if str(formula).startswith("=") else f"={formula}"
                elif entry.get("v") is not None:
                    cell.value = entry.get("v")
                style_ref = entry.get("s")
                # Univer may inline the style object on the cell, or point
                # at the styles table — handle both.
                pattern = None
                if isinstance(style_ref, dict):
                    pattern = (style_ref.get("n") or {}).get("pattern")
                elif style_ref:
                    pattern = ((styles.get(style_ref) or {}).get("n") or {}).get("pattern")
                if pattern:
                    cell.number_format = str(pattern)
        for merge in sheet.get("mergeData") or []:
            try:
                start = f"{get_column_letter(merge['startColumn'] + 1)}{merge['startRow'] + 1}"
                end = f"{get_column_letter(merge['endColumn'] + 1)}{merge['endRow'] + 1}"
                ws.merge_cells(f"{start}:{end}")
            except Exception:
                logger.warning("[documents] bad merge range skipped: %s", merge)
    wb.save(path)
