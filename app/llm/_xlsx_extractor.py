# FILE: app/llm/_xlsx_extractor.py
"""
Text extraction from Excel spreadsheets (.xlsx, .xls).

Converts spreadsheet content into a readable text representation
suitable for LLM context injection and RAG indexing.

Each sheet is rendered as a labelled section with tab-separated values.
Empty rows are skipped. Header rows (first row) are preserved.
"""
from __future__ import annotations

import io
import logging
from typing import Optional, Tuple

logger = logging.getLogger(__name__)

# Max cells to extract before truncating (prevents runaway on huge sheets)
MAX_CELLS = 50_000


def extract_xlsx_text(
    file_path: Optional[str] = None,
    file_bytes: Optional[bytes] = None,
) -> Tuple[str, Optional[str]]:
    """
    Extract text content from an Excel file.

    Args:
        file_path: Path to .xlsx/.xls file on disk.
        file_bytes: Raw file bytes (alternative to path).

    Returns:
        (text, error) — extracted text and optional error message.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "", "openpyxl not installed (pip install openpyxl)"

    try:
        if file_bytes:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        elif file_path:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        else:
            return "", "No file provided"
    except Exception as e:
        return "", f"Failed to open workbook: {e}"

    parts: list[str] = []
    cell_count = 0

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_lines: list[str] = []

            for row in ws.iter_rows(values_only=True):
                if cell_count >= MAX_CELLS:
                    sheet_lines.append(f"[TRUNCATED — {MAX_CELLS} cell limit reached]")
                    break

                # Skip completely empty rows
                values = [str(v) if v is not None else "" for v in row]
                if not any(values):
                    continue

                sheet_lines.append("\t".join(values))
                cell_count += len(values)

            if sheet_lines:
                parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(sheet_lines))
    finally:
        wb.close()

    if not parts:
        return "", "Workbook contains no readable data"

    return "\n\n".join(parts), None
